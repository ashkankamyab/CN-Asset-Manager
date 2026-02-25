import csv
import io

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q
from django.http import HttpResponse
from django.views import View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from assets.models import Asset, AssetRelationship


class ExportMixin:
    """Shared filtering logic for exports."""

    def get_filtered_queryset(self, request):
        qs = Asset.objects.exclude(status=Asset.Status.DECOMMISSIONED).select_related('aws_account', 'category').prefetch_related('outgoing_relationships__target_asset')
        search = request.GET.get('search')
        if search:
            qs = qs.filter(
                Q(name__icontains=search)
                | Q(asset_id__icontains=search)
                | Q(aws_resource_id__icontains=search)
            )
        if request.GET.get('asset_type'):
            qs = qs.filter(asset_type=request.GET['asset_type'])
        if request.GET.get('aws_service_type'):
            qs = qs.filter(aws_service_type=request.GET['aws_service_type'])
        if request.GET.get('criticality'):
            qs = qs.filter(criticality=request.GET['criticality'])
        if request.GET.get('status'):
            qs = qs.filter(status=request.GET['status'])
        if request.GET.get('aws_account'):
            qs = qs.filter(aws_account_id=request.GET['aws_account'])
        if request.GET.get('aws_region'):
            qs = qs.filter(aws_region=request.GET['aws_region'])

        # Exclude filters (comma-separated values)
        exclude_map = {
            'exclude_asset_type': 'asset_type__in',
            'exclude_aws_service_type': 'aws_service_type__in',
            'exclude_criticality': 'criticality__in',
            'exclude_status': 'status__in',
            'exclude_aws_account': 'aws_account_id__in',
            'exclude_aws_region': 'aws_region__in',
        }
        for param, lookup in exclude_map.items():
            raw = request.GET.get(param, '')
            values = [v.strip() for v in raw.split(',') if v.strip()]
            if values:
                qs = qs.exclude(**{lookup: values})

        return qs

    def _format_dependencies(self, asset):
        deps = []
        for rel in asset.outgoing_relationships.all():
            target = rel.target_asset
            deps.append(f'{rel.get_relationship_type_display()}: {target.asset_id} - {target.name}')
        return '; '.join(deps)

    def _format_datetime(self, dt):
        if not dt:
            return ''
        return dt.strftime('%Y-%m-%d %H:%M')

    def get_asset_row(self, asset):
        return [
            asset.asset_id,
            asset.name,
            asset.get_asset_type_display(),
            asset.get_aws_service_type_display() if asset.aws_service_type else '',
            str(asset.aws_account) if asset.aws_account else '',
            asset.aws_region,
            asset.aws_resource_id,
            asset.aws_resource_arn,
            asset.get_status_display(),
            asset.get_criticality_display(),
            asset.owner,
            asset.version,
            asset.get_data_classification_display(),
            'Yes' if asset.gdpr_relevant else 'No',
            'Yes' if asset.contains_personal_data else 'No',
            'Yes' if asset.backup_enabled else 'No',
            'Yes' if asset.monitoring_enabled else 'No',
            asset.vendor,
            asset.url,
            ', '.join(asset.ip_addresses) if asset.ip_addresses else '',
            ', '.join(asset.dns_names) if asset.dns_names else '',
            self._format_dependencies(asset),
            asset.description,
            asset.notes,
            self._format_datetime(asset.discovered_at),
        ]


HEADERS = [
    'Asset ID', 'Name', 'Type', 'AWS Service', 'AWS Account', 'Region',
    'Resource ID', 'ARN', 'Status', 'Criticality', 'Owner', 'Version',
    'Data Classification', 'GDPR Relevant', 'Personal Data',
    'Backup Enabled', 'Monitoring Enabled', 'Vendor', 'URL',
    'IP Addresses', 'DNS Names', 'Dependencies',
    'Description', 'Notes', 'Last Audit Date&Time',
]


class ExportAssetsCSVView(LoginRequiredMixin, ExportMixin, View):
    def get(self, request):
        assets = self.get_filtered_queryset(request)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="tisax_assets.csv"'
        writer = csv.writer(response)
        writer.writerow(HEADERS)
        for asset in assets:
            writer.writerow(self.get_asset_row(asset))
        return response


class ExportAssetsExcelView(LoginRequiredMixin, ExportMixin, View):
    def get(self, request):
        assets = list(self.get_filtered_queryset(request))
        wb = Workbook()

        # -- Summary sheet --
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')

        ws_summary.append(['Cloud Native Asset Manager - Export Summary'])
        ws_summary.append([])
        ws_summary.append(['Total Assets', len(assets)])
        by_service = {}
        by_account = {}
        by_criticality = {}
        for a in assets:
            svc = a.get_aws_service_type_display() if a.aws_service_type else 'Non-AWS'
            by_service[svc] = by_service.get(svc, 0) + 1
            acct = str(a.aws_account) if a.aws_account else 'No Account'
            by_account[acct] = by_account.get(acct, 0) + 1
            crit = a.get_criticality_display()
            by_criticality[crit] = by_criticality.get(crit, 0) + 1

        ws_summary.append([])
        ws_summary.append(['By AWS Service'])
        for svc, count in sorted(by_service.items(), key=lambda x: -x[1]):
            ws_summary.append([svc, count])
        ws_summary.append([])
        ws_summary.append(['By Account'])
        for acct, count in sorted(by_account.items(), key=lambda x: -x[1]):
            ws_summary.append([acct, count])
        ws_summary.append([])
        ws_summary.append(['By Criticality'])
        for crit, count in sorted(by_criticality.items()):
            ws_summary.append([crit, count])

        ws_summary.column_dimensions['A'].width = 40
        ws_summary.column_dimensions['B'].width = 15

        # -- By Service sheets --
        ws_by_service = wb.create_sheet('By Service')
        ws_by_service.append(HEADERS)
        for col_num, _ in enumerate(HEADERS, 1):
            cell = ws_by_service.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
        sorted_assets = sorted(assets, key=lambda a: (a.aws_service_type or 'ZZZ', a.name))
        for asset in sorted_assets:
            ws_by_service.append(self._get_asset_row(asset))

        # -- By Account sheet --
        ws_by_account = wb.create_sheet('By Account')
        ws_by_account.append(HEADERS)
        for col_num, _ in enumerate(HEADERS, 1):
            cell = ws_by_account.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
        sorted_assets = sorted(assets, key=lambda a: (str(a.aws_account or ''), a.name))
        for asset in sorted_assets:
            ws_by_account.append(self._get_asset_row(asset))

        # -- Full List sheet --
        ws_full = wb.create_sheet('Full List')
        ws_full.append(HEADERS)
        for col_num, _ in enumerate(HEADERS, 1):
            cell = ws_full.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
        for asset in assets:
            ws_full.append(self._get_asset_row(asset))

        # Auto-size columns for data sheets
        for ws in [ws_by_service, ws_by_account, ws_full]:
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="tisax_assets.xlsx"'
        return response

    def _get_asset_row(self, asset):
        return self.get_asset_row(asset)
